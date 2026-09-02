"""CSV and Excel export utilities — Python examples."""

import csv
import io
import zipfile
from pathlib import Path

import pandas as pd


def export_csv_pandas(rows: list[dict], path: str) -> None:
    """Small dataset: pandas to CSV."""
    df = pd.DataFrame(rows)
    df.to_csv(path, index=False)


def export_csv_streaming(cursor, path: str, headers: list[str]) -> None:
    """Large dataset: streaming CSV with a generator."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in cursor:
            writer.writerow(row)


def export_excel_multi_sheet(sheets: dict[str, list[dict]], path: str) -> None:
    """Excel with multiple sheets."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, rows in sheets.items():
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def export_csv_zip(datasets: dict[str, list[dict]], output_path: str) -> None:
    """Export multiple CSV files as a ZIP archive."""
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, rows in datasets.items():
            buffer = io.StringIO()
            if rows:
                writer = csv.DictWriter(buffer, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
            zf.writestr(f"{filename}.csv", buffer.getvalue())


def export_excel_with_formulas(path: str) -> None:
    """Export to Excel with formulas."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    wb.calculation.calcMode = "auto"
    wb.save(path)


if __name__ == "__main__":
    users = [
        {"id": 1, "name": "Alice", "email": "alice@example.com"},
        {"id": 2, "name": "Bob", "email": "bob@example.com"},
    ]

    export_csv_pandas(users, "users.csv")
    print("Wrote users.csv")

    export_excel_multi_sheet({"Users": users}, "report.xlsx")
    print("Wrote report.xlsx")

    export_csv_zip({"users": users, "admins": []}, "export.zip")
    print("Wrote export.zip")

    export_excel_with_formulas("formulas.xlsx")
    print("Wrote formulas.xlsx")
